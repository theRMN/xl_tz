import os

import openpyxl

from config import HOME_DIR, FOLDER_FILE_NAMES


def find_files():
    file_list = []

    # for all reports
    # for i in FOLDER_FILE_NAMES.keys():
    #     file_dir = os.path.join(HOME_DIR + '\\Desktop\\', i)
    #     file_ext = r'.xlsx'
    #     file_list += [f'{file_dir}\\{x}' for x in os.listdir(file_dir) if x.endswith(file_ext)]

    # for reports from "ГПХ" folder
    file_dir = os.path.join(HOME_DIR + '\\Desktop\\', 'ГПХ')
    file_ext = r'.xlsx'
    file_list += [f'{file_dir}\\{x}' for x in os.listdir(file_dir) if x.endswith(file_ext)]

    return file_list


def get_mapping():
    filename = 'mapping.xlsx'
    book = openpyxl.load_workbook(filename=filename, read_only=True)
    sheet = book.worksheets[0]
    mapping_dict = {}

    for row in range(1, sheet.max_row + 1):
        purchase_status = sheet[row][0].value
        agreement_status = sheet[row][1].value
        for_report = sheet[row][2].value
        mapping_dict[(purchase_status, agreement_status)] = for_report

    return mapping_dict


def modification_xl(mapping, file_list):
    for filename in file_list:
        book = openpyxl.load_workbook(filename=filename)
        sheet = book.worksheets[0]
        sheet.insert_cols(25)
        sheet['Y2'] = 'Статус для свода'

        for row in range(3, sheet.max_row + 1):
            agreement_status = sheet[row][23].value
            purchase_status = sheet[row][20].value
            key = (purchase_status, agreement_status)

            if key in mapping.keys():
                sheet[row][24].value = mapping[key]

        book.save(filename)


def run_xl():
    modification_xl(mapping=get_mapping(), file_list=find_files())


if __name__ == '__main__':
    run_xl()
